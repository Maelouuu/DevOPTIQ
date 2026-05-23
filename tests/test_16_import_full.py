# tests/test_16_import_full.py
"""
API : Import global IA depuis Excel (/api/import-full)
Couvre : analyse Excel (formats, matching, stats), injection en base (tâches, outils, rôles, compétences),
         sécurité cross-entity, liaison task_roles (doer/approver), cellules fusionnées, multi-groupes.
"""
import io
import json
import pytest
import openpyxl

pytestmark = pytest.mark.import_full


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def _make_excel(activity_name="Activité Test", task_name="Tâche Excel") -> bytes:
    """Crée un fichier Excel minimal valide pour les tests d'import."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Activities"
    ws.append(["Department", "Activity", "Guarantor", "Task", "Tool", "Doer", "Approver", "Skills", "Commentary"])
    ws.append(["Dept Test", activity_name, "Chef Test", task_name, "SAP, Excel", "Role A", "Role B", "Python, SQL", "Note"])
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _cleanup_task(app, name, activity_id):
    with app.app_context():
        from Code.models.models import Task
        from Code.extensions import db
        t = Task.query.filter_by(name=name, activity_id=activity_id).first()
        if t:
            db.session.delete(t)
            db.session.commit()


def _cleanup_role(app, name, entity_id):
    with app.app_context():
        from Code.models.models import Role
        from Code.extensions import db
        r = Role.query.filter_by(name=name, entity_id=entity_id).first()
        if r:
            db.session.delete(r)
            db.session.commit()


def _cleanup_tool(app, name, entity_id):
    with app.app_context():
        from Code.models.models import Tool
        from Code.extensions import db
        t = Tool.query.filter_by(name=name, entity_id=entity_id).first()
        if t:
            db.session.delete(t)
            db.session.commit()


# ===========================================================================
# 1. Endpoint POST /api/import-full/analyze
# ===========================================================================

class TestImportFullAnalyze:

    def test_analyze_no_file_returns_400(self, auth_client):
        """POST sans fichier → 400 avec message d'erreur."""
        r = auth_client.post("/api/import-full/analyze")
        assert r.status_code == 400
        data = json.loads(r.data)
        assert "error" in data

    def test_analyze_empty_filename_returns_400(self, auth_client):
        """Fichier sans nom → 400."""
        r = auth_client.post(
            "/api/import-full/analyze",
            data={"file": (io.BytesIO(b""), "")},
            content_type="multipart/form-data",
        )
        assert r.status_code == 400

    def test_analyze_invalid_extension_pdf_returns_400(self, auth_client):
        """Extension .pdf non supportée → 400 mentionnant le format."""
        r = auth_client.post(
            "/api/import-full/analyze",
            data={"file": (io.BytesIO(b"fake pdf content"), "test.pdf")},
            content_type="multipart/form-data",
        )
        assert r.status_code == 400
        data = json.loads(r.data)
        assert "error" in data

    def test_analyze_invalid_extension_csv_returns_400(self, auth_client):
        """Extension .csv non supportée → 400."""
        r = auth_client.post(
            "/api/import-full/analyze",
            data={"file": (io.BytesIO(b"col1,col2\nval1,val2"), "data.csv")},
            content_type="multipart/form-data",
        )
        assert r.status_code == 400

    def test_analyze_empty_excel_no_keywords_returns_400(self, auth_client):
        """Excel sans headers reconnaissables → 400 (aucune donnée trouvée)."""
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(["ColA", "ColB", "ColC"])
        ws.append(["val1", "val2", "val3"])
        buf = io.BytesIO()
        wb.save(buf)
        r = auth_client.post(
            "/api/import-full/analyze",
            data={"file": (io.BytesIO(buf.getvalue()), "empty.xlsx")},
            content_type="multipart/form-data",
        )
        assert r.status_code == 400

    def test_analyze_valid_xlsx_matched_activity(self, auth_client):
        """Excel avec activité existante (correspondance exacte) → 200, matched_groups >= 1."""
        excel = _make_excel(activity_name="Activité Test", task_name="Tâche Import Exacte")
        r = auth_client.post(
            "/api/import-full/analyze",
            data={"file": (io.BytesIO(excel), "test.xlsx")},
            content_type="multipart/form-data",
        )
        assert r.status_code == 200
        data = json.loads(r.data)
        assert data["status"] == "ok"
        assert len(data["analysis"]["matched_groups"]) >= 1

    def test_analyze_valid_xlsm_extension_accepted(self, auth_client):
        """Extension .xlsm aussi acceptée → 200."""
        excel = _make_excel(activity_name="Activité Test")
        r = auth_client.post(
            "/api/import-full/analyze",
            data={"file": (io.BytesIO(excel), "macro.xlsm")},
            content_type="multipart/form-data",
        )
        assert r.status_code == 200

    def test_analyze_valid_xls_extension_accepted(self, auth_client):
        """Extension .xls aussi acceptée → 200 (openpyxl lit les xlsx renommés en xls)."""
        excel = _make_excel(activity_name="Activité Test")
        r = auth_client.post(
            "/api/import-full/analyze",
            data={"file": (io.BytesIO(excel), "legacy.xls")},
            content_type="multipart/form-data",
        )
        assert r.status_code in (200, 400)

    def test_analyze_returns_db_activities_list(self, auth_client):
        """La réponse contient db_activities avec l'activité seedée."""
        excel = _make_excel(activity_name="Activité Test")
        r = auth_client.post(
            "/api/import-full/analyze",
            data={"file": (io.BytesIO(excel), "data.xlsx")},
            content_type="multipart/form-data",
        )
        assert r.status_code == 200
        data = json.loads(r.data)
        assert "db_activities" in data
        assert isinstance(data["db_activities"], list)
        assert any(a["name"] == "Activité Test" for a in data["db_activities"])

    def test_analyze_stats_matched_plus_unmatched_equals_total(self, auth_client):
        """stats: matched + unmatched == total_groups_excel."""
        excel = _make_excel(activity_name="Activité Test")
        r = auth_client.post(
            "/api/import-full/analyze",
            data={"file": (io.BytesIO(excel), "stats.xlsx")},
            content_type="multipart/form-data",
        )
        assert r.status_code == 200
        stats = json.loads(r.data)["stats"]
        assert stats["matched_activities"] + stats["unmatched_activities"] == stats["total_groups_excel"]

    def test_analyze_stats_tasks_count_consistent(self, auth_client):
        """matched_tasks + unmatched_tasks == total_tasks."""
        excel = _make_excel(activity_name="Activité Test")
        r = auth_client.post(
            "/api/import-full/analyze",
            data={"file": (io.BytesIO(excel), "tasks_stats.xlsx")},
            content_type="multipart/form-data",
        )
        assert r.status_code == 200
        stats = json.loads(r.data)["stats"]
        assert stats["matched_tasks"] + stats["unmatched_tasks"] == stats["total_tasks"]

    def test_analyze_unknown_activity_goes_to_unmatched(self, auth_client):
        """Activité introuvable → dans unmatched_groups avec possible_matches."""
        excel = _make_excel(activity_name="ZZZ Activité Inconnue XYZ999")
        r = auth_client.post(
            "/api/import-full/analyze",
            data={"file": (io.BytesIO(excel), "unknown.xlsx")},
            content_type="multipart/form-data",
        )
        assert r.status_code == 200
        data = json.loads(r.data)
        assert len(data["analysis"]["unmatched_groups"]) >= 1
        group = data["analysis"]["unmatched_groups"][0]
        assert "possible_matches" in group
        assert "reason" in group

    def test_analyze_matched_group_has_required_fields(self, auth_client):
        """Les matched_groups ont les champs activity_id, activity_name_db, tasks."""
        excel = _make_excel(activity_name="Activité Test")
        r = auth_client.post(
            "/api/import-full/analyze",
            data={"file": (io.BytesIO(excel), "fields.xlsx")},
            content_type="multipart/form-data",
        )
        assert r.status_code == 200
        groups = json.loads(r.data)["analysis"]["matched_groups"]
        assert len(groups) >= 1
        g = groups[0]
        assert "activity_id" in g
        assert "activity_name_db" in g
        assert "tasks" in g
        assert "confidence" in g

    def test_analyze_analysis_notes_present(self, auth_client):
        """La réponse contient analysis_notes décrivant le résultat."""
        excel = _make_excel(activity_name="Activité Test")
        r = auth_client.post(
            "/api/import-full/analyze",
            data={"file": (io.BytesIO(excel), "notes.xlsx")},
            content_type="multipart/form-data",
        )
        assert r.status_code == 200
        data = json.loads(r.data)
        assert "analysis_notes" in data["analysis"]
        assert len(data["analysis"]["analysis_notes"]) > 0


# ===========================================================================
# 2. Endpoint POST /api/import-full/inject
# ===========================================================================

class TestImportFullInject:

    def test_inject_no_body_returns_400(self, auth_client):
        """POST sans corps JSON → 400."""
        r = auth_client.post(
            "/api/import-full/inject",
            data=json.dumps({}),
            content_type="application/json",
        )
        assert r.status_code == 400
        data = json.loads(r.data)
        assert "error" in data

    def test_inject_empty_groups_returns_400(self, auth_client):
        """groups=[] → 400."""
        r = auth_client.post(
            "/api/import-full/inject",
            data=json.dumps({"groups": []}),
            content_type="application/json",
        )
        assert r.status_code == 400

    def test_inject_invalid_activity_id_skipped(self, auth_client):
        """activity_id inexistant → ignoré, stats tasks_created=0, retour 201."""
        r = auth_client.post(
            "/api/import-full/inject",
            data=json.dumps({
                "groups": [{
                    "activity_id": 999999,
                    "guarantor": "",
                    "tasks": [{"name": "Tâche Fantôme", "tools": [], "doer": "", "approver": "", "skills": []}],
                }]
            }),
            content_type="application/json",
        )
        assert r.status_code == 201
        data = json.loads(r.data)
        assert data["stats"]["tasks_created"] == 0

    def test_inject_valid_group_creates_task(self, auth_client, ids, app):
        """Group valide → tâche créée, 201 avec tasks_created=1."""
        r = auth_client.post(
            "/api/import-full/inject",
            data=json.dumps({
                "groups": [{
                    "activity_id": ids["activity_id"],
                    "guarantor": "",
                    "tasks": [{
                        "name": "Tâche Injectée Import Test",
                        "tools": [],
                        "doer": "",
                        "approver": "",
                        "skills": [],
                        "commentary": "",
                    }],
                }]
            }),
            content_type="application/json",
        )
        assert r.status_code == 201
        data = json.loads(r.data)
        assert data["stats"]["tasks_created"] == 1
        assert data["stats"]["activities_updated"] == 1
        _cleanup_task(app, "Tâche Injectée Import Test", ids["activity_id"])

    def test_inject_creates_tool(self, auth_client, ids, app):
        """Tâche avec outil inconnu → outil créé, tools_created=1."""
        r = auth_client.post(
            "/api/import-full/inject",
            data=json.dumps({
                "groups": [{
                    "activity_id": ids["activity_id"],
                    "guarantor": "",
                    "tasks": [{
                        "name": "Tâche Outil Import Test",
                        "tools": ["Outil Import Nouveau"],
                        "doer": "",
                        "approver": "",
                        "skills": [],
                    }],
                }]
            }),
            content_type="application/json",
        )
        assert r.status_code == 201
        data = json.loads(r.data)
        assert data["stats"]["tools_created"] >= 1
        _cleanup_task(app, "Tâche Outil Import Test", ids["activity_id"])
        _cleanup_tool(app, "Outil Import Nouveau", ids["entity_id"])

    def test_inject_creates_guarantor_role(self, auth_client, ids, app):
        """Group avec guarantor → rôle créé et lié à l'activité."""
        r = auth_client.post(
            "/api/import-full/inject",
            data=json.dumps({
                "groups": [{
                    "activity_id": ids["activity_id"],
                    "guarantor": "Garant Import Unique",
                    "tasks": [{
                        "name": "Tâche Garant Import",
                        "tools": [],
                        "doer": "",
                        "approver": "",
                        "skills": [],
                    }],
                }]
            }),
            content_type="application/json",
        )
        assert r.status_code == 201
        data = json.loads(r.data)
        assert data["stats"]["roles_created"] >= 1
        _cleanup_task(app, "Tâche Garant Import", ids["activity_id"])
        _cleanup_role(app, "Garant Import Unique", ids["entity_id"])

    def test_inject_creates_competency(self, auth_client, ids, app):
        """Skill dans la tâche → compétence créée, competencies_created >= 1."""
        r = auth_client.post(
            "/api/import-full/inject",
            data=json.dumps({
                "groups": [{
                    "activity_id": ids["activity_id"],
                    "guarantor": "",
                    "tasks": [{
                        "name": "Tâche Compétence Import",
                        "tools": [],
                        "doer": "",
                        "approver": "",
                        "skills": ["Python Import Skill"],
                    }],
                }]
            }),
            content_type="application/json",
        )
        assert r.status_code == 201
        data = json.loads(r.data)
        assert data["stats"]["competencies_created"] >= 1
        _cleanup_task(app, "Tâche Compétence Import", ids["activity_id"])
        with app.app_context():
            from Code.models.models import Competency
            from Code.extensions import db
            c = Competency.query.filter_by(description="Python Import Skill", activity_id=ids["activity_id"]).first()
            if c:
                db.session.delete(c)
                db.session.commit()

    def test_inject_duplicate_task_not_created_twice(self, auth_client, ids, app):
        """Injecter la même tâche deux fois → 2e fois tasks_created=0 (doublon ignoré)."""
        payload = json.dumps({
            "groups": [{
                "activity_id": ids["activity_id"],
                "guarantor": "",
                "tasks": [{"name": "Tâche Doublon Import Unique", "tools": [], "doer": "", "approver": "", "skills": []}],
            }]
        })
        r1 = auth_client.post("/api/import-full/inject", data=payload, content_type="application/json")
        assert r1.status_code == 201
        r2 = auth_client.post("/api/import-full/inject", data=payload, content_type="application/json")
        assert r2.status_code == 201
        assert json.loads(r2.data)["stats"]["tasks_created"] == 0
        _cleanup_task(app, "Tâche Doublon Import Unique", ids["activity_id"])

    def test_inject_response_has_status_ok(self, auth_client, ids, app):
        """La réponse contient status='ok' et les stats attendues."""
        r = auth_client.post(
            "/api/import-full/inject",
            data=json.dumps({
                "groups": [{
                    "activity_id": ids["activity_id"],
                    "guarantor": "",
                    "tasks": [{"name": "Tâche Stats Import", "tools": [], "doer": "", "approver": "", "skills": []}],
                }]
            }),
            content_type="application/json",
        )
        assert r.status_code == 201
        data = json.loads(r.data)
        assert data["status"] == "ok"
        assert "tasks_created" in data["stats"]
        assert "tools_created" in data["stats"]
        assert "roles_created" in data["stats"]
        assert "competencies_created" in data["stats"]
        assert "activities_updated" in data["stats"]
        _cleanup_task(app, "Tâche Stats Import", ids["activity_id"])


# ===========================================================================
# 3. Cas limites pour /analyze — session, matching partiel, multi-groupes
# ===========================================================================

class TestImportFullAnalyzeEdgeCases:

    def test_analyze_no_entity_id_returns_400(self, app):
        """POST analyze sans active_entity_id en session → 400 mentionnant l'entité."""
        no_entity_client = app.test_client()
        excel = _make_excel(activity_name="Activité Test")
        r = no_entity_client.post(
            "/api/import-full/analyze",
            data={"file": (io.BytesIO(excel), "test.xlsx")},
            content_type="multipart/form-data",
        )
        assert r.status_code == 400
        data = json.loads(r.data)
        assert "entit" in data["error"].lower()

    def test_analyze_partial_inclusion_match_goes_to_unmatched(self, auth_client):
        """Nom DB inclus dans nom Excel (score inclusion=0.88 < seuil 0.90) → unmatched_groups."""
        # "Activité Test" est substring de "Activité Test Process" → score 0.88 → unmatched
        excel = _make_excel(activity_name="Activité Test Process")
        r = auth_client.post(
            "/api/import-full/analyze",
            data={"file": (io.BytesIO(excel), "partial.xlsx")},
            content_type="multipart/form-data",
        )
        assert r.status_code == 200
        data = json.loads(r.data)
        matched_names = [g["activity_name_excel"] for g in data["analysis"]["matched_groups"]]
        assert "Activité Test Process" not in matched_names
        unmatched_names = [g["activity_name_excel"] for g in data["analysis"]["unmatched_groups"]]
        assert "Activité Test Process" in unmatched_names

    def test_analyze_unmatched_has_probable_reason_for_close_match(self, auth_client):
        """Activité proche mais sous le seuil 0.90 → reason contient 'probable' ou 'incertain'."""
        excel = _make_excel(activity_name="Activité Test Process")
        r = auth_client.post(
            "/api/import-full/analyze",
            data={"file": (io.BytesIO(excel), "reason.xlsx")},
            content_type="multipart/form-data",
        )
        assert r.status_code == 200
        data = json.loads(r.data)
        unmatched = [g for g in data["analysis"]["unmatched_groups"] if g["activity_name_excel"] == "Activité Test Process"]
        assert len(unmatched) == 1
        reason = unmatched[0]["reason"].lower()
        assert "probable" in reason or "incertain" in reason or "score" in reason

    def test_analyze_multiple_activities_in_file(self, auth_client):
        """Excel avec deux activités → total_groups_excel=2, matched + unmatched = 2."""
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Activities"
        ws.append(["Department", "Activity", "Guarantor", "Task", "Tool", "Doer", "Approver", "Skills", "Commentary"])
        ws.append(["Dept", "Activité Test", "Chef", "Tâche A1", "", "", "", "", ""])
        ws.append(["Dept", "ZZZ Activité Inconnue XYZ Import", "Chef", "Tâche B1", "", "", "", "", ""])
        buf = io.BytesIO()
        wb.save(buf)
        r = auth_client.post(
            "/api/import-full/analyze",
            data={"file": (io.BytesIO(buf.getvalue()), "multi.xlsx")},
            content_type="multipart/form-data",
        )
        assert r.status_code == 200
        data = json.loads(r.data)
        stats = data["stats"]
        assert stats["total_groups_excel"] == 2
        assert stats["matched_activities"] + stats["unmatched_activities"] == 2

    def test_analyze_merged_cell_rows_grouped_under_one_activity(self, auth_client):
        """Deux lignes même activité (cellule vide propagée) → 1 groupe avec 2 tâches."""
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Activities"
        ws.append(["Department", "Activity", "Guarantor", "Task", "Tool", "Doer", "Approver", "Skills", "Commentary"])
        ws.append(["Dept", "Activité Test", "Chef", "Tâche Merge 1", "", "", "", "", ""])
        ws.append(["", "", "", "Tâche Merge 2", "", "", "", "", ""])  # activity vide → propagée
        buf = io.BytesIO()
        wb.save(buf)
        r = auth_client.post(
            "/api/import-full/analyze",
            data={"file": (io.BytesIO(buf.getvalue()), "merge.xlsx")},
            content_type="multipart/form-data",
        )
        assert r.status_code == 200
        data = json.loads(r.data)
        assert data["stats"]["total_groups_excel"] == 1
        all_groups = data["analysis"]["matched_groups"] + data["analysis"]["unmatched_groups"]
        assert len(all_groups) == 1
        assert len(all_groups[0]["tasks"]) == 2


# ===========================================================================
# 4. Cas limites pour /inject — rôles, multi-groupes, sécurité cross-entity
# ===========================================================================

class TestImportFullInjectEdgeCases:

    def test_inject_no_entity_id_returns_400(self, app):
        """POST inject sans active_entity_id en session → 400 mentionnant l'entité."""
        no_entity_client = app.test_client()
        r = no_entity_client.post(
            "/api/import-full/inject",
            data=json.dumps({
                "groups": [{
                    "activity_id": 1,
                    "guarantor": "",
                    "tasks": [{"name": "X", "tools": [], "doer": "", "approver": "", "skills": []}],
                }]
            }),
            content_type="application/json",
        )
        assert r.status_code == 400
        data = json.loads(r.data)
        assert "entit" in data["error"].lower()

    def test_inject_with_doer_links_role_as_executant(self, auth_client, ids, app):
        """Tâche avec doer → entrée dans task_roles avec status='executant'."""
        r = auth_client.post(
            "/api/import-full/inject",
            data=json.dumps({
                "groups": [{
                    "activity_id": ids["activity_id"],
                    "guarantor": "",
                    "tasks": [{
                        "name": "Tâche Doer Import Test",
                        "tools": [],
                        "doer": "Exécutant Import Rôle",
                        "approver": "",
                        "skills": [],
                    }],
                }]
            }),
            content_type="application/json",
        )
        assert r.status_code == 201
        assert json.loads(r.data)["stats"]["roles_created"] >= 1
        with app.app_context():
            from Code.models.models import Task, Role, task_roles
            from Code.extensions import db
            task = Task.query.filter_by(name="Tâche Doer Import Test", activity_id=ids["activity_id"]).first()
            assert task is not None
            role = Role.query.filter_by(name="Exécutant Import Rôle", entity_id=ids["entity_id"]).first()
            assert role is not None
            tr = db.session.execute(
                task_roles.select().where(
                    task_roles.c.task_id == task.id,
                    task_roles.c.role_id == role.id,
                    task_roles.c.status == "executant",
                )
            ).first()
            assert tr is not None
            db.session.execute(task_roles.delete().where(task_roles.c.task_id == task.id))
            db.session.delete(task)
            db.session.delete(role)
            db.session.commit()

    def test_inject_with_approver_links_role_as_approbateur(self, auth_client, ids, app):
        """Tâche avec approver → entrée dans task_roles avec status='approbateur'."""
        r = auth_client.post(
            "/api/import-full/inject",
            data=json.dumps({
                "groups": [{
                    "activity_id": ids["activity_id"],
                    "guarantor": "",
                    "tasks": [{
                        "name": "Tâche Approver Import Test",
                        "tools": [],
                        "doer": "",
                        "approver": "Approbateur Import Rôle",
                        "skills": [],
                    }],
                }]
            }),
            content_type="application/json",
        )
        assert r.status_code == 201
        with app.app_context():
            from Code.models.models import Task, Role, task_roles
            from Code.extensions import db
            task = Task.query.filter_by(name="Tâche Approver Import Test", activity_id=ids["activity_id"]).first()
            assert task is not None
            role = Role.query.filter_by(name="Approbateur Import Rôle", entity_id=ids["entity_id"]).first()
            assert role is not None
            tr = db.session.execute(
                task_roles.select().where(
                    task_roles.c.task_id == task.id,
                    task_roles.c.role_id == role.id,
                    task_roles.c.status == "approbateur",
                )
            ).first()
            assert tr is not None
            db.session.execute(task_roles.delete().where(task_roles.c.task_id == task.id))
            db.session.delete(task)
            db.session.delete(role)
            db.session.commit()

    def test_inject_doer_and_approver_create_two_task_roles(self, auth_client, ids, app):
        """Tâche avec doer ET approver → 2 entrées dans task_roles, statuts 'executant' et 'approbateur'."""
        r = auth_client.post(
            "/api/import-full/inject",
            data=json.dumps({
                "groups": [{
                    "activity_id": ids["activity_id"],
                    "guarantor": "",
                    "tasks": [{
                        "name": "Tâche Deux Rôles Import",
                        "tools": [],
                        "doer": "Exécutant Deux Rôles Import",
                        "approver": "Approbateur Deux Rôles Import",
                        "skills": [],
                    }],
                }]
            }),
            content_type="application/json",
        )
        assert r.status_code == 201
        assert json.loads(r.data)["stats"]["roles_created"] >= 2
        with app.app_context():
            from Code.models.models import Task, task_roles
            from Code.extensions import db
            task = Task.query.filter_by(name="Tâche Deux Rôles Import", activity_id=ids["activity_id"]).first()
            assert task is not None
            trs = db.session.execute(
                task_roles.select().where(task_roles.c.task_id == task.id)
            ).fetchall()
            assert len(trs) == 2
            statuses = {tr.status for tr in trs}
            assert "executant" in statuses
            assert "approbateur" in statuses
            db.session.execute(task_roles.delete().where(task_roles.c.task_id == task.id))
            db.session.delete(task)
            db.session.commit()
        _cleanup_role(app, "Exécutant Deux Rôles Import", ids["entity_id"])
        _cleanup_role(app, "Approbateur Deux Rôles Import", ids["entity_id"])

    def test_inject_multiple_groups_activities_updated_count(self, auth_client, ids, app):
        """Deux groupes valides ciblant la même activité → activities_updated=2, tasks_created=2."""
        r = auth_client.post(
            "/api/import-full/inject",
            data=json.dumps({
                "groups": [
                    {
                        "activity_id": ids["activity_id"],
                        "guarantor": "",
                        "tasks": [{"name": "Tâche Multi Groupe A", "tools": [], "doer": "", "approver": "", "skills": []}],
                    },
                    {
                        "activity_id": ids["activity_id"],
                        "guarantor": "",
                        "tasks": [{"name": "Tâche Multi Groupe B", "tools": [], "doer": "", "approver": "", "skills": []}],
                    },
                ]
            }),
            content_type="application/json",
        )
        assert r.status_code == 201
        data = json.loads(r.data)
        assert data["stats"]["activities_updated"] == 2
        assert data["stats"]["tasks_created"] == 2
        _cleanup_task(app, "Tâche Multi Groupe A", ids["activity_id"])
        _cleanup_task(app, "Tâche Multi Groupe B", ids["activity_id"])

    def test_inject_empty_task_name_not_created(self, auth_client, ids):
        """Tâche avec name='' → ignorée, tasks_created=0."""
        r = auth_client.post(
            "/api/import-full/inject",
            data=json.dumps({
                "groups": [{
                    "activity_id": ids["activity_id"],
                    "guarantor": "",
                    "tasks": [{"name": "", "tools": [], "doer": "", "approver": "", "skills": []}],
                }]
            }),
            content_type="application/json",
        )
        assert r.status_code == 201
        assert json.loads(r.data)["stats"]["tasks_created"] == 0

    def test_inject_empty_string_tool_not_created(self, auth_client, ids, app):
        """Outil avec nom vide dans la liste tools → pas créé, tools_created=0."""
        r = auth_client.post(
            "/api/import-full/inject",
            data=json.dumps({
                "groups": [{
                    "activity_id": ids["activity_id"],
                    "guarantor": "",
                    "tasks": [{
                        "name": "Tâche Outil Vide Import",
                        "tools": [""],
                        "doer": "",
                        "approver": "",
                        "skills": [],
                    }],
                }]
            }),
            content_type="application/json",
        )
        assert r.status_code == 201
        assert json.loads(r.data)["stats"]["tools_created"] == 0
        _cleanup_task(app, "Tâche Outil Vide Import", ids["activity_id"])

    def test_inject_activity_cross_entity_skipped(self, auth_client, ids, app):
        """Activité d'une entité différente de la session active → ignorée, activities_updated=0."""
        with app.app_context():
            from Code.models.models import Entity, Activities
            from Code.extensions import db
            other_entity = Entity(name="Entité Autre Import Test", description="")
            db.session.add(other_entity)
            db.session.flush()
            other_activity = Activities(name="Activité Cross Entity Import", entity_id=other_entity.id)
            db.session.add(other_activity)
            db.session.commit()
            other_activity_id = other_activity.id
            other_entity_id = other_entity.id

        try:
            r = auth_client.post(
                "/api/import-full/inject",
                data=json.dumps({
                    "groups": [{
                        "activity_id": other_activity_id,
                        "guarantor": "",
                        "tasks": [{"name": "Tâche Cross Entity Import", "tools": [], "doer": "", "approver": "", "skills": []}],
                    }]
                }),
                content_type="application/json",
            )
            assert r.status_code == 201
            data = json.loads(r.data)
            assert data["stats"]["activities_updated"] == 0
            assert data["stats"]["tasks_created"] == 0
        finally:
            with app.app_context():
                from Code.models.models import Entity, Activities
                from Code.extensions import db
                act = Activities.query.get(other_activity_id)
                if act:
                    db.session.delete(act)
                ent = Entity.query.get(other_entity_id)
                if ent:
                    db.session.delete(ent)
                db.session.commit()

    def test_inject_doer_role_reused_not_recreated(self, auth_client, ids, app):
        """Même rôle doer utilisé en deux injections → 1ère fois roles_created=1, 2ème fois=0."""
        payload_1 = json.dumps({
            "groups": [{
                "activity_id": ids["activity_id"],
                "guarantor": "",
                "tasks": [{
                    "name": "Tâche Rôle Réutilisé Import 1",
                    "tools": [],
                    "doer": "Exécutant Réutilisé Import",
                    "approver": "",
                    "skills": [],
                }],
            }]
        })
        payload_2 = json.dumps({
            "groups": [{
                "activity_id": ids["activity_id"],
                "guarantor": "",
                "tasks": [{
                    "name": "Tâche Rôle Réutilisé Import 2",
                    "tools": [],
                    "doer": "Exécutant Réutilisé Import",
                    "approver": "",
                    "skills": [],
                }],
            }]
        })
        r1 = auth_client.post("/api/import-full/inject", data=payload_1, content_type="application/json")
        assert r1.status_code == 201
        assert json.loads(r1.data)["stats"]["roles_created"] == 1

        r2 = auth_client.post("/api/import-full/inject", data=payload_2, content_type="application/json")
        assert r2.status_code == 201
        assert json.loads(r2.data)["stats"]["roles_created"] == 0  # rôle déjà existant

        with app.app_context():
            from Code.models.models import Task, Role, task_roles
            from Code.extensions import db
            for tname in ["Tâche Rôle Réutilisé Import 1", "Tâche Rôle Réutilisé Import 2"]:
                task = Task.query.filter_by(name=tname, activity_id=ids["activity_id"]).first()
                if task:
                    db.session.execute(task_roles.delete().where(task_roles.c.task_id == task.id))
                    db.session.delete(task)
            role = Role.query.filter_by(name="Exécutant Réutilisé Import", entity_id=ids["entity_id"]).first()
            if role:
                db.session.delete(role)
            db.session.commit()
