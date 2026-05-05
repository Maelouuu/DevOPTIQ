# Code/routes/export.py
"""
Export des données d'une entité vers Excel ou HTML.
Upload-file  : sauvegarde le fichier en base de données (LargeBinary) — persistant sur cloud.
Serve-file   : sert le fichier depuis la DB (ou chemin local en dev).
"""
import os
import io
import mimetypes
from datetime import datetime

from flask import Blueprint, request, jsonify, send_file, Response
from werkzeug.utils import secure_filename
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from sqlalchemy import text as _sql_text

from Code.extensions import db
from Code.models.models import (
    Activities, Role, Savoir, SavoirFaire, Aptitude, Competency,
    activity_roles, Entity, FileBlob
)


def _get_role_mission(role_id):
    """Récupère mission_generale depuis la table roles via SQL brut."""
    row = db.session.execute(
        _sql_text("SELECT mission_generale FROM roles WHERE id = :rid"),
        {"rid": role_id}
    ).fetchone()
    return row[0] if row and row[0] else ""

export_bp = Blueprint("export", __name__)


# ──────────────────────────────────────────────
# Route : uploader un fichier → stocké en DB
# ──────────────────────────────────────────────
@export_bp.route("/utils/upload-file", methods=["POST"])
def upload_file():
    f = request.files.get("file")
    if not f or not f.filename:
        return jsonify({"error": "Aucun fichier fourni"}), 400

    filename = secure_filename(f.filename) or "fichier"
    mime = f.mimetype or mimetypes.guess_type(filename)[0] or "application/octet-stream"
    data = f.read()

    blob = FileBlob(filename=filename, mimetype=mime, data=data)
    db.session.add(blob)
    db.session.commit()

    return jsonify({
        "path": f"/utils/file/{blob.id}",
        "original_name": f.filename
    }), 201


# ──────────────────────────────────────────────
# Route : servir un fichier depuis la DB
# ──────────────────────────────────────────────
@export_bp.route("/utils/file/<int:file_id>")
def serve_db_file(file_id):
    blob = FileBlob.query.get(file_id)
    if not blob:
        return jsonify({"error": "Fichier introuvable"}), 404
    return Response(
        blob.data,
        mimetype=blob.mimetype,
        headers={"Content-Disposition": f'inline; filename="{blob.filename}"'}
    )


# ──────────────────────────────────────────────
# Route : serve-file (point d'entrée universel)
#   - /utils/file/123       → DB
#   - chemin absolu local   → filesystem (dev local uniquement)
# ──────────────────────────────────────────────
@export_bp.route("/utils/serve-file")
def serve_local_file():
    path = request.args.get("path", "").strip()
    if not path:
        return jsonify({"error": "Chemin manquant"}), 400

    # Fichier DB (nouvelle méthode)
    if path.startswith("/utils/file/"):
        try:
            file_id = int(path.rsplit("/", 1)[-1])
        except ValueError:
            return jsonify({"error": "ID de fichier invalide"}), 400
        blob = FileBlob.query.get(file_id)
        if not blob:
            return jsonify({"error": "Fichier introuvable en base"}), 404
        return Response(
            blob.data,
            mimetype=blob.mimetype,
            headers={"Content-Disposition": f'inline; filename="{blob.filename}"'}
        )

    # Fichier local absolu (dev local uniquement — ne fonctionne pas sur cloud)
    if os.path.exists(path):
        mime, _ = mimetypes.guess_type(path)
        return send_file(path, mimetype=mime or "application/octet-stream",
                         as_attachment=False, download_name=os.path.basename(path))

    return jsonify({"error": "Fichier introuvable (chemin local inaccessible sur cet environnement)"}), 404


# ──────────────────────────────────────────────
# Collecte des données (simplifié)
# ──────────────────────────────────────────────
def _collect_entity_data(entity_id, role_id=None):
    """
    Collecte les données pour l'export :
    - role_id fourni → activités Garant du rôle + savoirs/SF/aptitudes/compétences de ces activités
    - pas de role_id  → toutes les activités de l'entité + idem
    """
    entity = Entity.query.get(entity_id)
    if not entity:
        return None

    role = None
    if role_id:
        role = Role.query.get(role_id)

    # Activités
    if role_id:
        activities = (
            Activities.query
            .join(activity_roles, activity_roles.c.activity_id == Activities.id)
            .filter(
                activity_roles.c.role_id == role_id,
                activity_roles.c.status == 'Garant',
                Activities.entity_id == entity_id,
            )
            .order_by(Activities.name)
            .all()
        )
    else:
        activities = (
            Activities.query
            .filter_by(entity_id=entity_id)
            .order_by(Activities.name)
            .all()
        )

    # Agréger savoirs / SF / aptitudes / compétences de toutes ces activités
    all_savoirs, all_sf, all_aptitudes, all_competencies = [], [], [], []
    for act in activities:
        all_savoirs.extend(act.savoirs)
        all_sf.extend(act.savoir_faires)
        all_aptitudes.extend(act.aptitudes)
        all_competencies.extend(act.competencies)

    return {
        "entity": entity,
        "role": role,
        "mission_generale": _get_role_mission(role_id) if role_id else "",
        "activities": activities,
        "savoirs": all_savoirs,
        "savoir_faires": all_sf,
        "aptitudes": all_aptitudes,
        "competencies": all_competencies,
    }


# ──────────────────────────────────────────────
# Export Excel
# ──────────────────────────────────────────────
PURPLE     = "5B21B6"
PURPLE_MID = "7C3AED"
INDIGO     = "4338CA"
VIOLET_LT  = "EDE9FE"
VIOLET_MID = "DDD6FE"
WHITE      = "FFFFFF"
GRAY_LT    = "F5F3FF"

def _hdr_fill(hex_color): return PatternFill("solid", fgColor=hex_color)
def _border():
    s = Side(border_style="thin", color="C4B5FD")
    return Border(left=s, right=s, top=s, bottom=s)

def _make_excel(data, role_label=None):
    wb = Workbook()
    entity = data["entity"]
    role   = data.get("role")

    # ── Helpers ──────────────────────────────────────────────
    def _sheet_header(ws, headers, widths, fill_color):
        for i, (h, w) in enumerate(zip(headers, widths), 1):
            c = ws.cell(row=1, column=i, value=h)
            c.font = Font(bold=True, color=WHITE, size=11)
            c.fill = _hdr_fill(fill_color)
            c.alignment = Alignment(horizontal="center", vertical="center")
            c.border = _border()
            ws.column_dimensions[get_column_letter(i)].width = w
        ws.row_dimensions[1].height = 22
        ws.freeze_panes = "A2"

    def _add_rows(ws, rows, start=2):
        for ri, vals in enumerate(rows, start):
            fill = _hdr_fill(VIOLET_LT) if ri % 2 == 0 else _hdr_fill(WHITE)
            for ci, v in enumerate(vals, 1):
                c = ws.cell(row=ri, column=ci, value=v)
                c.fill = fill
                c.border = _border()
                c.alignment = Alignment(wrap_text=True, vertical="top")

    def write_kv(ws, row, key, value, col=PURPLE):
        k = ws.cell(row=row, column=1, value=key)
        k.font = Font(bold=True, color=col)
        k.fill = _hdr_fill(GRAY_LT)
        k.border = _border()
        v = ws.cell(row=row, column=2, value=str(value or ""))
        v.border = _border()
        v.alignment = Alignment(wrap_text=True)
        ws.row_dimensions[row].height = 18
        return row + 1

    def write_title(ws, row, title):
        c = ws.cell(row=row, column=1, value=title)
        c.font = Font(bold=True, color=WHITE, size=11)
        c.fill = _hdr_fill(PURPLE)
        c.alignment = Alignment(vertical="center", wrap_text=True)
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=2)
        ws.row_dimensions[row].height = 22
        return row + 1

    # ── Feuille 1 : Résumé ──────────────────────────────────
    ws = wb.active
    ws.title = "Résumé"
    ws.column_dimensions["A"].width = 26
    ws.column_dimensions["B"].width = 55

    r = 1
    r = write_title(ws, r, "Informations générales")
    r = write_kv(ws, r, "Entité", entity.name)
    r = write_kv(ws, r, "Date d'export", datetime.now().strftime("%d/%m/%Y %H:%M"))

    if role:
        r += 1
        r = write_title(ws, r, "Rôle exporté")
        r = write_kv(ws, r, "Nom du rôle", role.name)
        r = write_kv(ws, r, "Mission", data.get("mission_generale") or "—")
    else:
        r += 1
        r = write_title(ws, r, "Périmètre")
        r = write_kv(ws, r, "Filtre", "Toutes les activités de l'entité")

    r += 1
    r = write_title(ws, r, "Statistiques")
    r = write_kv(ws, r, "Activités", len(data["activities"]))
    r = write_kv(ws, r, "Savoirs", len(data["savoirs"]))
    r = write_kv(ws, r, "Savoir-faires", len(data["savoir_faires"]))
    r = write_kv(ws, r, "HSC / Aptitudes", len(data["aptitudes"]))
    r = write_kv(ws, r, "Compétences", len(data["competencies"]))

    # ── Feuille 2 : Activités (noms uniquement) ──────────────
    ws2 = wb.create_sheet("Activités")
    _sheet_header(ws2, ["#", "Nom de l'activité"], [6, 70], PURPLE)
    _add_rows(ws2, [(i, act.name) for i, act in enumerate(data["activities"], 1)])

    # ── Feuille 3 : Savoirs ──────────────────────────────────
    ws3 = wb.create_sheet("Savoirs")
    _sheet_header(ws3, ["Activité", "Savoir"], [35, 70], INDIGO)
    _add_rows(ws3, [(s.activity.name if s.activity else "—", s.description) for s in data["savoirs"]])

    # ── Feuille 4 : Savoir-faires ────────────────────────────
    ws4 = wb.create_sheet("Savoir-faires")
    _sheet_header(ws4, ["Activité", "Savoir-faire"], [35, 70], PURPLE_MID)
    _add_rows(ws4, [(sf.activity.name if sf.activity else "—", sf.description) for sf in data["savoir_faires"]])

    # ── Feuille 5 : HSC / Aptitudes ─────────────────────────
    ws5 = wb.create_sheet("HSC - Aptitudes")
    _sheet_header(ws5, ["Activité", "Aptitude / HSC"], [35, 70], PURPLE)
    _add_rows(ws5, [(a.activity.name if a.activity else "—", a.description) for a in data["aptitudes"]])

    # ── Feuille 6 : Compétences ──────────────────────────────
    ws6 = wb.create_sheet("Compétences")
    _sheet_header(ws6, ["Activité", "Compétence"], [35, 70], INDIGO)
    _add_rows(ws6, [(c.activity.name if c.activity else "—", c.description) for c in data["competencies"]])

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


# ──────────────────────────────────────────────
# Export HTML
# ──────────────────────────────────────────────
def _make_html(data, role_label=None):
    entity = data["entity"]
    role   = data.get("role")
    now    = datetime.now().strftime("%d/%m/%Y à %H:%M")

    def esc(s):
        if not s:
            return "—"
        return str(s).replace("&", "&amp;").replace("<", "&lt;").replace(">", "&gt;")

    def _list_section(title, items, color):
        if not items:
            return f'<p class="no-items">Aucun élément.</p>'
        rows = "".join(
            f'<tr><td class="act-ref">{esc(getattr(item, "activity", None) and item.activity.name or "—")}</td>'
            f'<td>{esc(item.description)}</td></tr>'
            for item in items
        )
        return (
            f'<div class="table-wrap">'
            f'<table><thead style="background:{color};">'
            f'<tr><th>Activité</th><th>Description</th></tr>'
            f'</thead><tbody>{rows}</tbody></table></div>'
        )

    # ── Section rôle (si export par rôle) ───────────────────
    role_section = ""
    if role:
        mission_text = data.get("mission_generale") or ""
        mission = esc(mission_text) if mission_text else "<em>Non renseignée</em>"
        role_section = f"""
        <div class="role-card">
          <div class="role-header">
            <div class="role-icon">👤</div>
            <div>
              <div class="role-name">{esc(role.name)}</div>
              <div class="role-label">Rôle exporté</div>
            </div>
          </div>
          <div class="role-mission">
            <span class="field-label">Mission</span>
            <p>{mission}</p>
          </div>
        </div>"""

    # ── Liste des activités ──────────────────────────────────
    accents = ["#6d28d9", "#4338ca", "#7c3aed", "#5b21b6"]
    acts_items = "".join(
        f'<li style="border-left-color:{accents[i % 4]};">'
        f'<span class="act-num">{i+1:02d}</span>'
        f'<span class="act-name">{esc(act.name)}</span>'
        f'</li>'
        for i, act in enumerate(data["activities"])
    ) or '<li class="no-items">Aucune activité.</li>'

    # ── Sections savoirs / SF / aptitudes / compétences ─────
    savoirs_html = _list_section("Savoirs", data["savoirs"], "linear-gradient(to right,#4c1d95,#4338ca)")
    sf_html      = _list_section("Savoir-faires", data["savoir_faires"], "linear-gradient(to right,#5b21b6,#7c3aed)")
    apt_html     = _list_section("HSC / Aptitudes", data["aptitudes"], "linear-gradient(to right,#4c1d95,#4338ca)")
    comp_html    = _list_section("Compétences", data["competencies"], "linear-gradient(to right,#5b21b6,#7c3aed)")

    nb_acts = len(data["activities"])
    scope_label = f"Rôle : {esc(role.name)}" if role else "Toutes les activités"

    html = f"""<!DOCTYPE html>
<html lang="fr">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width, initial-scale=1.0">
<title>Export — {esc(entity.name)}</title>
<style>
  :root {{
    --purple: #5b21b6; --indigo: #4338ca; --violet: #7c3aed;
    --violet-lt: #ede9fe; --violet-mid: #ddd6fe;
    --white: #ffffff; --gray: #f5f3ff; --text: #1e1040; --muted: #6b7280;
  }}
  * {{ box-sizing: border-box; margin: 0; padding: 0; }}
  body {{ font-family: 'Segoe UI', system-ui, sans-serif; background: #f0ebff; color: var(--text); }}

  /* ── Hero ── */
  .hero {{
    background: linear-gradient(155deg, #12053d 0%, #2e0f6b 35%, #4c1d95 65%, #4338ca 100%);
    padding: 48px 40px 40px; position: relative; overflow: hidden;
  }}
  .hero::before {{
    content:''; position:absolute; top:-60px; right:-60px; width:260px; height:260px;
    background:radial-gradient(circle, rgba(216,180,254,.4) 0%, transparent 70%); border-radius:50%;
  }}
  .hero-brand {{ display:flex; align-items:center; gap:10px; margin-bottom:20px; }}
  .hero-logo {{ width:38px;height:38px;background:rgba(255,255,255,.15);border:1.5px solid rgba(255,255,255,.3);
    border-radius:10px;display:flex;align-items:center;justify-content:center;font-size:18px;color:#e9d5ff;font-weight:900; }}
  .hero-name {{ font-size:14px;font-weight:800;color:rgba(255,255,255,.8);letter-spacing:2px;text-transform:uppercase; }}
  .hero-title {{ font-size:30px;font-weight:800;color:#fff;line-height:1.2;margin-bottom:8px; }}
  .hero-sub {{ font-size:14px;color:rgba(221,214,254,.75); }}
  .hero-meta {{ display:flex;flex-wrap:wrap;gap:10px;margin-top:18px; }}
  .meta-pill {{ background:rgba(255,255,255,.1);border:1px solid rgba(196,181,253,.35);border-radius:20px;
    padding:5px 14px;font-size:12px;color:#e9d5ff;font-weight:600; }}
  .scope-pill {{ background:rgba(74,222,128,.15);border:1px solid rgba(74,222,128,.4);border-radius:20px;
    padding:5px 14px;font-size:12px;color:#86efac;font-weight:600; }}

  /* ── Contenu ── */
  .container {{ max-width:900px;margin:0 auto;padding:36px 24px; }}

  .section-title {{
    font-size:18px;font-weight:800;color:var(--purple);margin:32px 0 14px;
    display:flex;align-items:center;gap:10px;
  }}
  .section-title::after {{ content:'';flex:1;height:2px;background:linear-gradient(to right,var(--violet-mid),transparent); }}

  /* ── Carte rôle ── */
  .role-card {{
    background:#fff;border-radius:16px;border:1px solid var(--violet-mid);
    border-left:4px solid var(--purple);
    box-shadow:0 4px 16px rgba(109,40,217,.08);margin-bottom:28px;overflow:hidden;
  }}
  .role-header {{
    display:flex;align-items:center;gap:14px;
    padding:16px 24px;
    background:linear-gradient(to right,var(--gray),#fff);
  }}
  .role-icon {{ font-size:26px; }}
  .role-name {{ font-size:20px;font-weight:800;color:var(--text); }}
  .role-label {{ font-size:11px;font-weight:600;text-transform:uppercase;letter-spacing:.8px;color:var(--muted);margin-top:2px; }}
  .role-mission {{ padding:14px 24px 18px; }}
  .field-label {{ font-size:11px;font-weight:700;text-transform:uppercase;letter-spacing:.8px;color:var(--muted);display:block;margin-bottom:6px; }}
  .role-mission p {{ font-size:13.5px;color:var(--text);line-height:1.65; }}

  /* ── Liste activités ── */
  .act-list {{
    list-style:none;display:flex;flex-direction:column;gap:6px;
    background:#fff;border-radius:14px;border:1px solid var(--violet-mid);
    padding:14px;box-shadow:0 2px 10px rgba(109,40,217,.06);
  }}
  .act-list li {{
    display:flex;align-items:center;gap:12px;
    padding:9px 12px;border-radius:10px;
    border-left:3px solid var(--purple);
    background:var(--gray);
    transition:transform .15s;
  }}
  .act-list li:hover {{ transform:translateX(2px); }}
  .act-num {{
    width:28px;height:28px;border-radius:8px;background:var(--purple);
    color:#fff;font-size:11px;font-weight:800;
    display:flex;align-items:center;justify-content:center;flex-shrink:0;
  }}
  .act-name {{ font-size:13.5px;font-weight:600;color:var(--text); }}

  /* ── Tables compétences ── */
  .table-wrap {{ overflow-x:auto;border-radius:12px;border:1px solid var(--violet-mid);
    box-shadow:0 2px 10px rgba(109,40,217,.06); }}
  table {{ width:100%;border-collapse:collapse;font-size:13px; }}
  thead th {{ color:#fff;padding:10px 14px;text-align:left;font-weight:700;font-size:12px; }}
  tbody tr:nth-child(even) {{ background:var(--gray); }}
  tbody tr:nth-child(odd) {{ background:#fff; }}
  tbody td {{ padding:9px 14px;border-top:1px solid var(--violet-lt);vertical-align:top;line-height:1.55; }}
  .act-ref {{ color:var(--muted);font-size:12px;width:30%;font-style:italic; }}
  .no-items {{ color:var(--muted);font-style:italic;padding:12px 0; }}

  /* ── Footer ── */
  .footer {{ text-align:center;padding:32px;color:var(--muted);font-size:12px; }}
  .footer strong {{ color:var(--purple); }}

  @media print {{
    body {{ background:#fff; }}
    .hero {{ -webkit-print-color-adjust:exact;print-color-adjust:exact; }}
  }}
</style>
</head>
<body>

<div class="hero">
  <div class="hero-brand">
    <div class="hero-logo">◉</div>
    <span class="hero-name">OPTIQ</span>
  </div>
  <h1 class="hero-title">{esc(entity.name)}</h1>
  <p class="hero-sub">Export des données métier</p>
  <div class="hero-meta">
    <span class="meta-pill">📋 {nb_acts} activité{"s" if nb_acts != 1 else ""}</span>
    <span class="meta-pill">🗓 Exporté le {now}</span>
    <span class="scope-pill">🎯 {scope_label}</span>
  </div>
</div>

<div class="container">

  {role_section}

  <h2 class="section-title">Activités</h2>
  <ul class="act-list">{acts_items}</ul>

  <h2 class="section-title">Savoirs</h2>
  {savoirs_html}

  <h2 class="section-title">Savoir-faires</h2>
  {sf_html}

  <h2 class="section-title">HSC / Aptitudes</h2>
  {apt_html}

  <h2 class="section-title">Compétences</h2>
  {comp_html}

</div>

<div class="footer">
  Généré par <strong>OPTIQ</strong> · {now}
</div>

</body>
</html>"""

    return html


# ──────────────────────────────────────────────
# Routes d'export
# ──────────────────────────────────────────────
@export_bp.route("/export/entity")
def export_entity():
    entity_id = request.args.get("entity_id", type=int)
    role_id   = request.args.get("role_id",   type=int) or None
    fmt       = request.args.get("format", "excel").lower()

    if not entity_id:
        return jsonify({"error": "entity_id requis"}), 400

    data = _collect_entity_data(entity_id, role_id)
    if not data:
        return jsonify({"error": "Entité introuvable"}), 404

    role_label = data["role"].name if data.get("role") else None

    entity_name = data["entity"].name
    date_str    = datetime.now().strftime("%Y%m%d")
    role_suffix = f"_{role_label.replace(' ', '_')}" if role_label else ""

    if fmt == "html":
        html_content = _make_html(data, role_label)
        filename = f"export_{entity_name}{role_suffix}_{date_str}.html"
        return Response(
            html_content,
            mimetype="text/html",
            headers={"Content-Disposition": f'attachment; filename="{filename}"'}
        )
    else:
        buf = _make_excel(data, role_label)
        filename = f"export_{entity_name}{role_suffix}_{date_str}.xlsx"
        return send_file(
            buf, mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            as_attachment=True, download_name=filename
        )
