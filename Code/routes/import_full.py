# Code/routes/import_full.py
# Blueprint Flask — Import global IA depuis Excel
# Matching algorithmique (difflib) + OpenAI optionnel en enrichissement

import io
import os
import json
from difflib import SequenceMatcher

import openpyxl
from flask import Blueprint, request, jsonify, session
from sqlalchemy import func

from Code.extensions import db
from Code.models.models import (
    Activities, Task, Tool, Role, Competency,
    Entity, activity_roles, task_roles,
)

import_full_bp = Blueprint('import_full', __name__, url_prefix='/api/import-full')


# ---------------------------------------------------------------------------
# Lecture du fichier Excel — parse robuste avec merged cells propagées
# ---------------------------------------------------------------------------

def _parse_excel_bytes(data: bytes) -> list:
    """
    Lit le fichier Excel et retourne une liste de groupes par activité.
    Gère les merged cells en propagant les valeurs manquantes.
    """
    wb = openpyxl.load_workbook(io.BytesIO(data), data_only=True)

    # Chercher la feuille principale
    target_sheet = None
    for name in wb.sheetnames:
        if 'activity' in name.lower() or 'activit' in name.lower():
            target_sheet = wb[name]
            break
    if not target_sheet:
        target_sheet = wb.active

    ws = target_sheet

    # Identifier les headers (première ligne non vide avec des mots-clés reconnus)
    header_row = None
    header_row_num = None
    for i, row in enumerate(ws.iter_rows(min_row=1, max_row=10, values_only=True), 1):
        row_s = [str(v or '').strip().lower() for v in row]
        if any(kw in ' '.join(row_s) for kw in ('task', 'tool', 'activity', 'semi finish', 'department', 'doer')):
            header_row = row_s
            header_row_num = i
            break

    if not header_row:
        return []

    # Mapping colonnes → index
    col_map = {}
    kw_map = {
        'id':         ['id'],
        'department': ['department', 'dept'],
        'activity':   ['semi finish', 'semi-finish', 'activity', 'activit'],
        'guarantor':  ['guarantor', 'garant'],
        'task':       ['task', 'tâche'],
        'tool':       ['tool', 'outil'],
        'doer':       ['doer', 'executor'],
        'approver':   ['approver', 'checker', 'approbateur'],
        'skills':     ['skills', 'knowledge', 'competenc', 'savoir'],
        'commentary': ['comment', 'commentaire', 'note'],
    }
    for col_idx, header in enumerate(header_row):
        for key, kws in kw_map.items():
            if any(kw in header for kw in kws) and key not in col_map:
                col_map[key] = col_idx

    def _get(row, key, default=''):
        idx = col_map.get(key)
        if idx is None or idx >= len(row):
            return default
        val = row[idx]
        return str(val).strip() if val is not None else default

    def _split_csv(s: str) -> list:
        return [x.strip() for x in s.replace(';', ',').split(',') if x.strip()]

    # Lire les lignes de données (après le header)
    groups = []
    current_group = None
    last_activity = ''
    last_department = ''
    last_guarantor = ''

    for row in ws.iter_rows(min_row=header_row_num + 1, max_row=ws.max_row, values_only=True):
        row = list(row)

        activity_name = _get(row, 'activity')
        department = _get(row, 'department')
        guarantor = _get(row, 'guarantor')
        task_name = _get(row, 'task')
        tools_raw = _get(row, 'tool')
        doer = _get(row, 'doer')
        approver = _get(row, 'approver')
        skills_raw = _get(row, 'skills')
        commentary = _get(row, 'commentary')

        # Propager les merged cells
        if activity_name:
            last_activity = activity_name
        else:
            activity_name = last_activity
        if department:
            last_department = department
        else:
            department = last_department
        if guarantor:
            last_guarantor = guarantor
        else:
            guarantor = last_guarantor

        if not activity_name:
            continue

        # Nouvelle activité ou continuation
        if current_group is None or current_group['activity_name'] != activity_name:
            current_group = {
                'activity_name': activity_name,
                'department': department,
                'guarantor': guarantor,
                'tasks': [],
            }
            groups.append(current_group)

        if task_name:
            current_group['tasks'].append({
                'name': task_name,
                'tools': _split_csv(tools_raw),
                'doer': doer,
                'approver': approver,
                'skills': _split_csv(skills_raw),
                'commentary': commentary,
            })
        elif tools_raw and current_group['tasks']:
            # Outils supplémentaires sur la ligne suivante (merged cells)
            for t in _split_csv(tools_raw):
                if t not in current_group['tasks'][-1]['tools']:
                    current_group['tasks'][-1]['tools'].append(t)

    return [g for g in groups if g['tasks']]


# ---------------------------------------------------------------------------
# Matching algorithmique (aucune dépendance externe)
# ---------------------------------------------------------------------------

def _normalize(s: str) -> str:
    return s.strip().lower()


def _similarity(a: str, b: str) -> float:
    return SequenceMatcher(None, _normalize(a), _normalize(b)).ratio()


def _algorithmic_match(excel_groups: list, db_activities: list) -> dict:
    """
    Matching pur algorithmique :
    1. Correspondance exacte (casse ignorée)
    2. Correspondance par inclusion (l'un contient l'autre)
    3. Correspondance fuzzy via SequenceMatcher (seuil 0.60)
    """
    matched_groups = []
    unmatched_groups = []

    for group in excel_groups:
        excel_name = group['activity_name']
        excel_norm = _normalize(excel_name)

        best_act = None
        best_score = 0.0
        best_reason = ''

        for act in db_activities:
            db_norm = _normalize(act['name'])

            # 1. Exact
            if excel_norm == db_norm:
                best_act = act
                best_score = 1.0
                best_reason = 'Correspondance exacte'
                break

            # 2. Inclusion
            if excel_norm in db_norm or db_norm in excel_norm:
                score = 0.88
                if score > best_score:
                    best_act = act
                    best_score = score
                    best_reason = 'Correspondance partielle (contenu dans l\'autre)'
                continue

            # 3. Fuzzy
            score = _similarity(excel_name, act['name'])
            if score > best_score:
                best_act = act
                best_score = score
                best_reason = f'Correspondance approximative ({score:.0%})'

        if best_act and best_score >= 0.60:
            conf = 'high' if best_score >= 0.90 else ('medium' if best_score >= 0.75 else 'low')
            matched_groups.append({
                'activity_name_excel': excel_name,
                'activity_id': best_act['id'],
                'activity_name_db': best_act['name'],
                'confidence': conf,
                'match_reason': best_reason,
                'guarantor': group.get('guarantor', ''),
                'tasks': group['tasks'],
            })
        else:
            # Top 3 suggestions pour les non-matchés
            scored = sorted(
                db_activities,
                key=lambda a: _similarity(excel_name, a['name']),
                reverse=True,
            )
            possible = [
                {
                    'activity_id': a['id'],
                    'activity_name': a['name'],
                    'similarity': 'medium' if _similarity(excel_name, a['name']) >= 0.5 else 'low',
                }
                for a in scored[:3]
            ]
            reason = (
                f'Meilleur score : {best_score:.0%} — aucune correspondance fiable.'
                if best_act else 'Aucune activité dans la base.'
            )
            unmatched_groups.append({
                'activity_name_excel': excel_name,
                'reason': reason,
                'possible_matches': possible,
                'guarantor': group.get('guarantor', ''),
                'tasks': group['tasks'],
            })

    notes = (
        f'Analyse terminée : {len(matched_groups)} activité(s) mappée(s), '
        f'{len(unmatched_groups)} à résoudre manuellement.'
    )
    return {
        'matched_groups': matched_groups,
        'unmatched_groups': unmatched_groups,
        'analysis_notes': notes,
    }


# ---------------------------------------------------------------------------
# Enrichissement OpenAI optionnel (uniquement pour les non-matchés)
# ---------------------------------------------------------------------------

ENRICH_PROMPT = """\
Tu es un assistant d'import OPTIQ.
On t'envoie des groupes d'activités Excel qui N'ONT PAS pu être mappés automatiquement,
et la liste des activités de la base.

Pour chaque groupe non-mappé, essaie de trouver l'activité la plus proche dans la base
(matching sémantique, synonymes, traductions FR/EN).

Réponds UNIQUEMENT en JSON valide :
{
  "resolved": [
    {
      "activity_name_excel": "...",
      "activity_id": 42,
      "activity_name_db": "...",
      "confidence": "high"|"medium"|"low",
      "match_reason": "..."
    }
  ],
  "still_unmatched": ["nom excel", ...]
}
"""


def _try_openai_enrich(unmatched_groups: list, db_activities: list) -> dict | None:
    """Tente un enrichissement IA pour les non-matchés. Retourne None si indisponible."""
    api_key = os.getenv('OPENAI_API_KEY')
    if not api_key or not unmatched_groups:
        return None

    try:
        from openai import OpenAI
        payload = {
            'unmatched': [
                {'activity_name_excel': g['activity_name_excel']}
                for g in unmatched_groups
            ],
            'db_activities': db_activities,
        }
        client = OpenAI()
        model = os.getenv('OPENAI_CHATBOT_MODEL', 'gpt-4o-mini')
        resp = client.chat.completions.create(
            model=model,
            messages=[
                {'role': 'system', 'content': ENRICH_PROMPT},
                {'role': 'user', 'content': json.dumps(payload, ensure_ascii=False)},
            ],
            response_format={'type': 'json_object'},
            temperature=0.1,
            max_tokens=800,
        )
        return json.loads(resp.choices[0].message.content)
    except Exception as e:
        print(f'[ImportFull] OpenAI enrichissement ignoré : {e}')
        return None


# ---------------------------------------------------------------------------
# Endpoint : analyse
# ---------------------------------------------------------------------------

@import_full_bp.post('/analyze')
def analyze_excel():
    """
    Reçoit un fichier Excel, analyse algorithmiquement les activités,
    enrichit optionnellement avec OpenAI pour les non-matchés.
    """
    if 'file' not in request.files:
        return jsonify({'error': 'Aucun fichier fourni'}), 400

    file = request.files['file']
    if not file.filename:
        return jsonify({'error': 'Fichier vide'}), 400

    allowed = ('.xlsx', '.xls', '.xlsm')
    if not any(file.filename.lower().endswith(e) for e in allowed):
        return jsonify({'error': 'Format non supporté — utilisez .xlsx, .xls ou .xlsm'}), 400

    entity_id = session.get('active_entity_id')
    if not entity_id:
        return jsonify({
            'error': 'Aucune entité active — activez une entité dans la cartographie.'
        }), 400

    # Parse Excel
    try:
        excel_groups = _parse_excel_bytes(file.read())
    except Exception as e:
        return jsonify({'error': f'Erreur lecture Excel : {str(e)}'}), 400

    if not excel_groups:
        return jsonify({'error': 'Aucune donnée trouvée dans le fichier'}), 400

    # Activités en base
    activities = Activities.query.filter_by(entity_id=entity_id).order_by(Activities.name).all()
    db_activities = [{'id': a.id, 'name': a.name} for a in activities]

    if not db_activities:
        return jsonify({
            'error': 'Aucune activité dans cette entité. Importez d\'abord votre cartographie SVG.'
        }), 400

    # 1. Matching algorithmique (toujours)
    analysis = _algorithmic_match(excel_groups, db_activities)

    # 2. Enrichissement IA optionnel pour les non-matchés (silencieux si quota dépassé)
    if analysis['unmatched_groups']:
        ai_result = _try_openai_enrich(analysis['unmatched_groups'], db_activities)
        if ai_result:
            resolved_names = {r['activity_name_excel'] for r in ai_result.get('resolved', [])}
            resolved_map = {r['activity_name_excel']: r for r in ai_result.get('resolved', [])}

            still_unmatched = []
            for grp in analysis['unmatched_groups']:
                name = grp['activity_name_excel']
                if name in resolved_map:
                    r = resolved_map[name]
                    analysis['matched_groups'].append({
                        'activity_name_excel': name,
                        'activity_id': r['activity_id'],
                        'activity_name_db': r['activity_name_db'],
                        'confidence': r.get('confidence', 'medium'),
                        'match_reason': r.get('match_reason', 'Résolu par IA'),
                        'guarantor': grp.get('guarantor', ''),
                        'tasks': grp['tasks'],
                    })
                else:
                    still_unmatched.append(grp)

            analysis['unmatched_groups'] = still_unmatched
            analysis['analysis_notes'] = (
                f'Analyse hybride (algo + IA) : {len(analysis["matched_groups"])} mappée(s), '
                f'{len(analysis["unmatched_groups"])} à résoudre.'
            )

    # Statistiques
    matched = analysis['matched_groups']
    unmatched = analysis['unmatched_groups']
    total_tasks = sum(len(g['tasks']) for g in matched + unmatched)
    matched_tasks = sum(len(g['tasks']) for g in matched)

    return jsonify({
        'status': 'ok',
        'analysis': analysis,
        'stats': {
            'total_groups_excel': len(excel_groups),
            'matched_activities': len(matched),
            'unmatched_activities': len(unmatched),
            'total_tasks': total_tasks,
            'matched_tasks': matched_tasks,
            'unmatched_tasks': total_tasks - matched_tasks,
        },
        'db_activities': db_activities,
    })


# ---------------------------------------------------------------------------
# Endpoint : injection
# ---------------------------------------------------------------------------

@import_full_bp.post('/inject')
def inject_full():
    """
    Reçoit les groupes validés et les injecte en base.
    """
    data = request.get_json(force=True) or {}
    groups = data.get('groups', [])

    if not groups:
        return jsonify({'error': 'Aucun groupe à injecter'}), 400

    entity_id = session.get('active_entity_id')
    if not entity_id:
        return jsonify({'error': 'Aucune entité active'}), 400

    stats = {
        'tasks_created': 0,
        'tools_created': 0,
        'roles_created': 0,
        'competencies_created': 0,
        'activities_updated': 0,
    }

    try:
        for group in groups:
            activity_id = group.get('activity_id')
            if not activity_id:
                continue

            activity = Activities.query.get(activity_id)
            if not activity or activity.entity_id != entity_id:
                continue

            guarantor_name = (group.get('guarantor') or '').strip()

            # ── Garant ───────────────────────────────────────────────────
            if guarantor_name:
                role = _get_or_create_role(guarantor_name, entity_id, stats)
                _link_role_to_activity(role, activity, 'garant')

            # ── Tâches ───────────────────────────────────────────────────
            max_order = (
                db.session.query(func.max(Task.order))
                .filter_by(activity_id=activity_id)
                .scalar() or 0
            )

            for i, task_in in enumerate(group.get('tasks', [])):
                task_name = (task_in.get('name') or '').strip()
                if not task_name:
                    continue

                task = Task(
                    name=task_name,
                    description=task_in.get('commentary', '') or '',
                    order=max_order + i + 1,
                    activity_id=activity_id,
                )
                db.session.add(task)
                db.session.flush()
                stats['tasks_created'] += 1

                # Outils
                for tool_name in (task_in.get('tools') or []):
                    tool_name = tool_name.strip()
                    if not tool_name:
                        continue
                    tool = _get_or_create_tool(tool_name, entity_id, stats)
                    if tool not in task.tools:
                        task.tools.append(tool)

                # Doer
                doer_name = (task_in.get('doer') or '').strip()
                if doer_name:
                    doer_role = _get_or_create_role(doer_name, entity_id, stats)
                    _link_role_to_task(doer_role, task, 'executant')

                # Approbateur
                approver_name = (task_in.get('approver') or '').strip()
                if approver_name:
                    approver_role = _get_or_create_role(approver_name, entity_id, stats)
                    _link_role_to_task(approver_role, task, 'approbateur')

                # Compétences
                for skill in (task_in.get('skills') or []):
                    skill = skill.strip()
                    if not skill:
                        continue
                    exists = Competency.query.filter_by(
                        activity_id=activity_id,
                        description=skill,
                    ).first()
                    if not exists:
                        db.session.add(Competency(activity_id=activity_id, description=skill))
                        stats['competencies_created'] += 1

            stats['activities_updated'] += 1

        db.session.commit()
        return jsonify({'status': 'ok', 'stats': stats}), 201

    except Exception as e:
        db.session.rollback()
        return jsonify({'error': str(e)}), 500


# ---------------------------------------------------------------------------
# Helpers DB
# ---------------------------------------------------------------------------

def _get_or_create_tool(name: str, entity_id: int, stats: dict) -> Tool:
    tool = Tool.query.filter(
        Tool.entity_id == entity_id,
        func.lower(Tool.name) == name.lower(),
    ).first()
    if not tool:
        tool = Tool(name=name, entity_id=entity_id)
        db.session.add(tool)
        db.session.flush()
        stats['tools_created'] += 1
    return tool


def _get_or_create_role(name: str, entity_id: int, stats: dict) -> Role:
    role = Role.query.filter(
        Role.entity_id == entity_id,
        func.lower(Role.name) == name.lower(),
    ).first()
    if not role:
        role = Role(name=name, entity_id=entity_id)
        db.session.add(role)
        db.session.flush()
        stats['roles_created'] += 1
    return role


def _link_role_to_activity(role: Role, activity: Activities, status: str):
    exists = db.session.execute(
        activity_roles.select().where(
            activity_roles.c.activity_id == activity.id,
            activity_roles.c.role_id == role.id,
        )
    ).first()
    if not exists:
        db.session.execute(
            activity_roles.insert().values(
                activity_id=activity.id,
                role_id=role.id,
                status=status,
            )
        )
        db.session.flush()


def _link_role_to_task(role: Role, task: Task, status: str):
    exists = db.session.execute(
        task_roles.select().where(
            task_roles.c.task_id == task.id,
            task_roles.c.role_id == role.id,
        )
    ).first()
    if not exists:
        db.session.execute(
            task_roles.insert().values(
                task_id=task.id,
                role_id=role.id,
                status=status,
            )
        )
        db.session.flush()
